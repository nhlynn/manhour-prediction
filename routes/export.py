"""Export route blueprint for MHES.

Handles data export to Excel format matching the simple_resource template.
"""

import logging
import os
import re

from flask import Blueprint, current_app, jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

export_bp = Blueprint("export", __name__)


@export_bp.route("/excel", methods=["POST"])
def export_excel():
    """Export preview data to an Excel file in simple_resource folder.

    Expects JSON body with ``projectName`` and ``categories``.
    Returns the generated file as a download.

    As a side effect, if the project contains any Category, Task, or
    Activity Detail not already present in the knowledge base, the
    project is also saved into the knowledge base (named after the
    project) and embedded, so it becomes searchable via the chatbot.
    This never blocks or fails the download itself.
    """
    data = request.get_json(silent=True) or {}
    project_name = (data.get("projectName") or "").strip()
    categories = data.get("categories", [])

    if not project_name:
        return jsonify({"error": "Project name is required."}), 400

    if not categories:
        return jsonify({"error": "No data to export."}), 400

    # Sanitize filename
    safe_name = re.sub(r'[\\/*?:"<>|]', "_", project_name)
    filename = f"{safe_name}_manhour.xlsx"
    output_folder = os.path.join(current_app.root_path, "exports")
    os.makedirs(output_folder, exist_ok=True)
    filepath = os.path.join(output_folder, filename)

    try:
        _build_workbook(filepath, project_name, categories)
    except Exception as e:
        logger.error(f"Export failed: {e}")
        return jsonify({"error": str(e)}), 500

    _add_to_knowledge_base_if_new(safe_name, project_name, categories)

    return send_file(filepath, as_attachment=True, download_name=filename)


# ------------------------------------------------------------------
# Auto-add new knowledge to the KB + embeddings on export
# ------------------------------------------------------------------

def _add_to_knowledge_base_if_new(
    safe_name: str, project_name: str, categories: list
) -> None:
    """If the project has any new Category/Task/Activity, add it to the KB.

    Best-effort: any failure here is logged and swallowed so it never
    breaks the actual Excel download response.
    """
    try:
        from services.embedding_service import EmbeddingService

        emb_svc = EmbeddingService(
            model_name=current_app.config["EMBEDDING_MODEL"],
            embeddings_folder=current_app.config["EMBEDDINGS_FOLDER"],
        )
        existing_cats, existing_tasks, existing_details = _load_existing_kb_index(emb_svc)

        if not _has_new_items(categories, existing_cats, existing_tasks, existing_details):
            return

        kb_folder = current_app.config["KB_FOLDER"]
        os.makedirs(kb_folder, exist_ok=True)
        kb_filename = _unique_kb_filename(kb_folder, f"{safe_name}.xlsx")
        kb_filepath = os.path.join(kb_folder, kb_filename)

        _build_kb_ingest_workbook(kb_filepath, categories)
        emb_svc.process_excel_file(kb_filepath)

        logger.info(
            f"Project '{project_name}' contains new knowledge — "
            f"added to KB as '{kb_filename}' and embedded."
        )
    except Exception:
        logger.exception(
            f"Failed to add exported project '{project_name}' to the "
            f"knowledge base/embeddings; export itself is unaffected."
        )


def _load_existing_kb_index(emb_svc) -> tuple[set, set, set]:
    """Load the set of Category/Task/Activity names already in the KB.

    Scans every currently-embedded KB file's mapping JSON (not the raw
    Excel — the mapping is what the chatbot actually searches over).

    Returns:
        Tuple of (categories, (category, task) pairs, (category, task,
        detail) triples), all lowercased/stripped for comparison.
    """
    import json

    existing_cats: set[str] = set()
    existing_tasks: set[tuple[str, str]] = set()
    existing_details: set[tuple[str, str, str]] = set()

    metadata = emb_svc._load_metadata()
    for file_meta in metadata.values():
        mapping_path = file_meta.get("mapping_path", "")
        if not mapping_path or not os.path.isfile(mapping_path):
            continue
        with open(mapping_path, "r", encoding="utf-8") as f:
            nested_json = json.load(f)

        for cat in nested_json:
            cat_name = cat.get("category", "").strip().lower()
            if not cat_name:
                continue
            existing_cats.add(cat_name)
            for task in cat.get("tasks", []):
                task_name = task.get("task", "").strip().lower()
                existing_tasks.add((cat_name, task_name))
                for detail in task.get("task_details", []):
                    detail_name = detail.get("task_detail", "").strip().lower()
                    existing_details.add((cat_name, task_name, detail_name))

    return existing_cats, existing_tasks, existing_details


def _has_new_items(
    categories: list,
    existing_cats: set,
    existing_tasks: set,
    existing_details: set,
) -> bool:
    """Return True if any Category/Task/Activity isn't already in the KB."""
    for cat in categories:
        cat_name = (cat.get("category") or "").strip().lower()
        if not cat_name:
            continue
        if cat_name not in existing_cats:
            return True
        for task in cat.get("tasks", []):
            task_name = (task.get("task") or "").strip().lower()
            if (cat_name, task_name) not in existing_tasks:
                return True
            for act in task.get("activities", []):
                detail_name = (act.get("task_detail") or "").strip().lower()
                if (cat_name, task_name, detail_name) not in existing_details:
                    return True
    return False


def _unique_kb_filename(kb_folder: str, filename: str) -> str:
    """Return a filename guaranteed not to collide within kb_folder."""
    if not os.path.isfile(os.path.join(kb_folder, filename)):
        return filename

    from datetime import datetime

    base, ext = os.path.splitext(filename)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    candidate = f"{base}_{timestamp}{ext}"
    counter = 1
    while os.path.isfile(os.path.join(kb_folder, candidate)):
        candidate = f"{base}_{timestamp}_{counter}{ext}"
        counter += 1
    return candidate


def _build_kb_ingest_workbook(filepath: str, categories: list) -> None:
    """Build a KB-ingestible workbook (one row per Activity Detail).

    Uses the column layout excel_parser expects: Category, Task List,
    Activity Details, Estimate (Hours), Buffer (Hours) — so it can be
    embedded the same way as any manually-uploaded KB file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Man Hour"

    headers = ["Category", "Task List", "Activity Details", "Estimate (Hours)", "Buffer (Hours)"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    row = 2
    for cat in categories:
        cat_name = cat.get("category", "")
        for task in cat.get("tasks", []):
            task_name = task.get("task", "")
            buffer_hours = task.get("buffer_hours", 0)
            activities = task.get("activities", [])

            for i, act in enumerate(activities):
                ws.cell(row=row, column=1, value=cat_name)
                ws.cell(row=row, column=2, value=task_name)
                ws.cell(row=row, column=3, value=act.get("task_detail", ""))
                ws.cell(row=row, column=4, value=act.get("estimate_hours", 0))
                # Buffer is task-level — record it once, on the first
                # activity row, matching the convention used elsewhere
                # in hand-authored KB files (see excel_parser._map_columns).
                ws.cell(row=row, column=5, value=buffer_hours if i == 0 else 0)
                row += 1

    wb.save(filepath)


def _build_workbook(filepath: str, project_name: str, categories: list) -> None:
    """Build an Excel workbook matching the reference template format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Manhour"

    # --- Styles ---
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    cat_font = Font(bold=True)
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_align = Alignment(vertical="center", wrap_text=True)

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 15

    # --- Row 1-2: Title (merged) ---
    ws.merge_cells("A1:D2")
    title_cell = ws["A1"]
    title_cell.value = f"{project_name} Manhour"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Row 3: Empty ---

    # --- Row 4: Headers ---
    headers = ["Category", "Task List", "Estimate (Hours)", "Working Day"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Data rows ---
    row = 5
    grand_total = 0

    for cat in categories:
        cat_start_row = row
        cat_name = cat.get("category", "")

        # Each task as a numbered row (no activity detail flattening)
        task_num = 1
        cat_total_hours = 0

        for task in cat.get("tasks", []):
            task_name = task.get("task", "")
            total_hours = task.get("total_hours", 0)

            # Task List column: numbered task name
            ws.cell(row=row, column=2, value=f"{task_num}. {task_name}").alignment = wrap_align
            ws.cell(row=row, column=2).border = thin_border

            # Estimate (Hours) per task
            ws.cell(row=row, column=3, value=total_hours).alignment = center_align
            ws.cell(row=row, column=3).border = thin_border

            # Working Day per task (= hours / 8)
            ws.cell(row=row, column=4, value=f"=C{row}/8").alignment = center_align
            ws.cell(row=row, column=4).border = thin_border

            cat_total_hours += total_hours
            task_num += 1
            row += 1

        cat_end_row = row - 1
        if cat_end_row < cat_start_row:
            continue

        cat_row_count = cat_end_row - cat_start_row + 1

        # Category column (merged)
        if cat_row_count > 1:
            ws.merge_cells(
                start_row=cat_start_row, start_column=1,
                end_row=cat_end_row, end_column=1,
            )
        cat_cell = ws.cell(row=cat_start_row, column=1, value=cat_name)
        cat_cell.font = cat_font
        cat_cell.alignment = Alignment(vertical="center")
        cat_cell.border = thin_border
        for r in range(cat_start_row, cat_end_row + 1):
            ws.cell(row=r, column=1).border = thin_border

        grand_total += cat_total_hours

    # --- Total row ---
    total_row = row
    ws.cell(row=total_row, column=1, value="Total").font = total_font
    for col_idx in range(1, 5):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.fill = total_fill
        cell.border = thin_border
        cell.font = total_font

    ws.cell(row=total_row, column=3, value=grand_total).alignment = center_align
    ws.cell(row=total_row, column=4, value=f"=C{total_row}/8").alignment = center_align

    wb.save(filepath)
