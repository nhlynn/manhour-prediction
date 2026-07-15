"""Upload route blueprint for MHES.

Handles Excel file upload, duplicate detection, and knowledge base management.
Routes are kept thin — all logic lives in the service layer.
"""

import logging
import os

from flask import (
    Blueprint,
    current_app,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from services.excel_service import ExcelService
from services.embedding_service import EmbeddingService

logger = logging.getLogger(__name__)

upload_bp = Blueprint("upload", __name__)


# ------------------------------------------------------------------
# Service helpers
# ------------------------------------------------------------------

def _excel_service() -> ExcelService:
    return ExcelService(kb_folder=current_app.config["KB_FOLDER"])


def _embedding_service() -> EmbeddingService:
    return EmbeddingService(
        model_name=current_app.config["EMBEDDING_MODEL"],
        embeddings_folder=current_app.config["EMBEDDINGS_FOLDER"],
    )


# ------------------------------------------------------------------
# Routes
# ------------------------------------------------------------------

@upload_bp.route("/", methods=["GET"])
def upload_page() -> str:
    """Render the upload page with the list of imported files."""
    svc = _excel_service()
    emb = _embedding_service()
    kb_files = svc.list_knowledge_files()
    # Enrich each file entry with embedding info from central metadata
    for f in kb_files:
        f["has_embeddings"] = emb.has_index(f["filename"])
        emb_meta = emb.get_file_metadata(f["filename"])
        if emb_meta:
            f["num_categories"] = emb_meta.get("num_categories", 0)
            f["num_vectors"] = emb_meta.get("num_vectors", 0)
        else:
            f["num_categories"] = 0
            f["num_vectors"] = 0
    return render_template("upload.html", kb_files=kb_files)


@upload_bp.route("/template", methods=["GET"])
def download_template():
    """Download a blank knowledge-file template with the expected columns.

    The columns match what ``services.excel_parser`` looks for
    (flexibly, by keyword), so a file filled in from this template is
    guaranteed to be parsed and embedded correctly on upload.
    """
    filename = "MHES_KB_Template.xlsx"
    filepath = os.path.join(current_app.instance_path, filename)
    os.makedirs(current_app.instance_path, exist_ok=True)

    _build_template_workbook(filepath)

    return send_file(filepath, as_attachment=True, download_name=filename)


def _build_template_workbook(filepath: str) -> None:
    """Write a sample knowledge-file workbook with headers and example rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Man Hour"

    headers = ["Category", "Task List", "Activity Details", "Estimate (Hours)", "Buffer (Hours)"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Example rows: one task with three activities, and a second task,
    # showing that Category/Task List only need to be filled on the
    # first row of their group (merged cells or blank repeats both work,
    # since the parser forward-fills them).
    example_rows = [
        ["Wordpress Server", "Documentation", "Prepare installation", 0.5, 2],
        ["", "", "configuration", 0.5, ""],
        ["", "", "and handover documentation", 1, ""],
        ["Wordpress Server", "Deployment", "Deploy to production server", 2, 1],
        ["", "", "Smoke test key pages", 1, ""],
    ]
    start_row = 2
    last_row = start_row + len(example_rows) - 1
    for row_idx, row_data in enumerate(example_rows, start=start_row):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Thin border around every cell in the table, header included.
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = border

    # Merge Category / Task List / Buffer (Hours) across each task group
    # (a group starts wherever Task List is filled in), same convention
    # as the reference template.
    middle_align = Alignment(vertical="center")
    group_starts = [
        r for r in range(start_row, last_row + 1)
        if ws.cell(row=r, column=2).value not in (None, "")
    ]
    group_starts.append(last_row + 1)
    for i in range(len(group_starts) - 1):
        g_start, g_end = group_starts[i], group_starts[i + 1] - 1
        for col_idx in (1, 2, 5):
            ws.cell(row=g_start, column=col_idx).alignment = middle_align
            if g_end > g_start:
                ws.merge_cells(start_row=g_start, start_column=col_idx, end_row=g_end, end_column=col_idx)

    widths = {"A": 22, "B": 20, "C": 40, "D": 18, "E": 16}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(filepath)


@upload_bp.route("/check-duplicates", methods=["POST"])
def check_duplicates() -> tuple:
    """Check which of the selected filenames already exist in kb_knowledge.

    Expects a JSON body: ``{"filenames": ["a.xlsx", "b.xlsx"]}``.

    Returns:
        JSON with ``{"duplicates": ["a.xlsx"]}`` (only the ones that exist).
    """
    from werkzeug.utils import secure_filename as _secure

    data = request.get_json(silent=True) or {}
    filenames = data.get("filenames", [])
    svc = _excel_service()
    duplicates = [
        name for name in filenames if svc.file_exists(_secure(name))
    ]
    return jsonify({"duplicates": duplicates})


@upload_bp.route("/", methods=["POST"])
def upload_files() -> str:
    """Handle one or multiple Excel file uploads.

    Form fields:
        ``files``: One or more file inputs.
        ``duplicate_action``: ``"rename"`` (default) or ``"overwrite"``.

    After each successful save the embedding service is called automatically.
    """
    files = request.files.getlist("files")
    duplicate_action = request.form.get("duplicate_action", "rename")

    if not files or all(f.filename in (None, "") for f in files):
        flash("No files selected.", "warning")
        return redirect(url_for("upload.upload_page"))

    svc = _excel_service()
    emb = _embedding_service()
    success_count = 0
    fail_count = 0

    for file in files:
        if file.filename is None or file.filename.strip() == "":
            continue

        # Validate extension
        if not ExcelService.is_valid_extension(file.filename):
            flash(f"Skipped '{file.filename}': invalid file type.", "danger")
            fail_count += 1
            continue

        try:
            # Save file to kb_knowledge
            meta = svc.save_file(file, duplicate_action=duplicate_action)
            label = meta["filename"]
            if meta["overwritten"]:
                # Delete old embeddings before rebuilding
                emb.delete_index(label)
                flash(f"Overwritten: {label} ({meta['size_kb']} KB)", "success")
            else:
                flash(f"Uploaded: {label} ({meta['size_kb']} KB)", "success")

            # Auto-generate embeddings (reads all sheets)
            try:
                kb_path = os.path.join(current_app.config["KB_FOLDER"], meta["filename"])
                result = emb.process_excel_file(kb_path)
                flash(
                    f"Embeddings ready for {label}: "
                    f"{result['num_vectors']} text chunks from "
                    f"{result['num_categories']} category(ies).",
                    "info",
                )
            except Exception as e:
                logger.error(f"Embedding failed for '{label}': {e}")
                flash(
                    f"Uploaded '{label}' but embedding failed: {e}",
                    "warning",
                )

            success_count += 1

        except Exception as e:
            logger.error(f"Upload failed for '{file.filename}': {e}")
            flash(f"Failed to upload '{file.filename}': {e}", "danger")
            fail_count += 1

    if success_count:
        logger.info(f"Upload batch: {success_count} succeeded, {fail_count} failed")

    return redirect(url_for("upload.upload_page"))


@upload_bp.route("/delete/<filename>", methods=["POST"])
def delete_file(filename: str) -> str:
    """Delete a knowledge base file and its embeddings.

    Args:
        filename: Name of the file to remove.
    """
    svc = _excel_service()
    emb = _embedding_service()

    try:
        emb.delete_index(filename)
        if svc.delete_file(filename):
            flash(f"Deleted: {filename}", "success")
        else:
            flash(f"File not found: {filename}", "warning")
    except Exception as e:
        logger.error(f"Delete failed for '{filename}': {e}")
        flash(f"Delete failed: {e}", "danger")

    return redirect(url_for("upload.upload_page"))


@upload_bp.route("/reembed/<filename>", methods=["POST"])
def reembed_file(filename: str) -> str:
    """Re-generate embeddings for an existing knowledge base file.

    Args:
        filename: Name of the Excel file in kb_knowledge.
    """
    emb = _embedding_service()
    kb_path = os.path.join(current_app.config["KB_FOLDER"], filename)

    try:
        result = emb.process_excel_file(kb_path)
        flash(
            f"Embeddings regenerated for '{filename}': "
            f"{result['num_vectors']} vectors.",
            "info",
        )
    except Exception as e:
        logger.error(f"Re-embedding failed for '{filename}': {e}")
        flash(f"Re-embedding failed: {e}", "danger")

    return redirect(url_for("upload.upload_page"))
